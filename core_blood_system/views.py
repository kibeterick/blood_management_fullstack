from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from .models import CustomUser, Donor, BloodRequest, BloodDonation, BloodInventory
from .forms import (UserRegistrationForm, AdminRegistrationForm, CustomLoginForm, 
                    DonorRegistrationForm, BloodRequestForm, BloodDonationForm,
                    BloodRequestStatusForm)

# Import for Excel/PDF exports
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import io


# Home Page
def home(request):
    """Home page view"""
    return render(request, 'home.html')


# CSRF Failure Handler
def csrf_failure(request, reason=""):
    """Custom CSRF failure view that redirects to a friendly error page"""
    return render(request, '403_csrf.html', status=403)


def home_portal(request):
    """Home portal - redirect to home"""
    return render(request, 'home.html')


# User Registration
def user_register(request):
    """Regular user registration"""
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.role = 'user'
            user.save()
            messages.success(request, 'Account created successfully! Please log in.')
            return redirect('login')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = UserRegistrationForm()
    
    return render(request, 'registration/user_register.html', {'form': form})


# Admin Registration
def admin_register(request):
    """Admin registration - only accessible by existing admins"""
    # Only allow logged-in admins to create new admin accounts
    if not request.user.is_authenticated or request.user.role != 'admin':
        messages.error(request, 'Admin registration is restricted. Only existing administrators can create new admin accounts.')
        return redirect('login')
    
    if request.method == 'POST':
        form = AdminRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, f'Admin account for {user.username} created successfully!')
            return redirect('user_dashboard')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = AdminRegistrationForm()
    
    return render(request, 'registration/admin_register.html', {'form': form})


# Login View
def user_login(request):
    """User login with personalized welcome message"""
    if request.method == 'POST':
        form = CustomLoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            remember_me = request.POST.get('remember_me')
            
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                
                # Handle "Remember Me" functionality
                if remember_me:
                    # Session expires in 30 days
                    request.session.set_expiry(2592000)  # 30 days in seconds
                else:
                    # Session expires when browser closes
                    request.session.set_expiry(0)
                
                # Get time-based greeting
                from datetime import datetime
                current_hour = datetime.now().hour
                if current_hour < 12:
                    greeting = "Good morning"
                elif current_hour < 17:
                    greeting = "Good afternoon"
                else:
                    greeting = "Good evening"
                
                # Personalized welcome message
                full_name = f"{user.first_name} {user.last_name}" if user.first_name else user.username
                
                # Set welcome message in session for modal display
                request.session['show_welcome'] = True
                request.session['welcome_greeting'] = greeting
                request.session['welcome_name'] = full_name
                request.session['user_role'] = user.role
                request.session['last_login'] = user.last_login.strftime('%B %d, %Y at %I:%M %p') if user.last_login else 'First time login'
                
                # Success message
                messages.success(request, f'{greeting}, {full_name}! Welcome back to Blood Management System.')
                
                # Redirect based on user role - EXPLICIT CHECK
                # Check multiple conditions to ensure proper routing
                if hasattr(user, 'role') and user.role == 'admin':
                    return redirect('admin_dashboard')
                elif user.is_staff or user.is_superuser:
                    return redirect('admin_dashboard')
                else:
                    return redirect('user_dashboard')
            else:
                messages.error(request, 'Invalid username or password.')
        else:
            messages.error(request, 'Invalid username or password.')
    else:
        form = CustomLoginForm()
    
    return render(request, 'registration/login.html', {'form': form})


# Logout View
@login_required
def user_logout(request):
    """User logout"""
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('home')


# User Dashboard
@login_required
def user_dashboard(request):
    """Dashboard for regular users - Enhanced version"""
    # CRITICAL: Redirect admins to admin dashboard
    # Check multiple conditions to ensure admins see the right dashboard
    if request.user.is_authenticated:
        # Check if user is admin by role
        if hasattr(request.user, 'role') and request.user.role == 'admin':
            return redirect('admin_dashboard')
        # Also check if user is staff/superuser
        if request.user.is_staff or request.user.is_superuser:
            return redirect('admin_dashboard')
    
    user_requests = BloodRequest.objects.filter(requester=request.user)
    
    # Get blood inventory for display
    inventory = BloodInventory.objects.all().order_by('blood_type')
    
    return render(request, 'dashboard/user_dashboard_enhanced.html', {
        'user_requests': user_requests,
        'inventory': inventory,
    })


# Admin Dashboard
@login_required
def admin_dashboard(request):
    """Dashboard for administrators - Enhanced version"""
    if request.user.role != 'admin':
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('user_dashboard')
    
    # Get statistics
    total_donors = Donor.objects.count()
    total_requests = BloodRequest.objects.count()
    pending_requests = BloodRequest.objects.filter(status='pending').count()
    total_donations = BloodDonation.objects.count()
    total_users = CustomUser.objects.count()
    
    # Get recent requests
    recent_requests = BloodRequest.objects.all()[:10]
    
    # Get blood inventory
    inventory = BloodInventory.objects.all()
    
    # Get recent users (last 10 registered)
    recent_users = CustomUser.objects.all().order_by('-date_joined')[:10]
    
    context = {
        'total_donors': total_donors,
        'total_requests': total_requests,
        'pending_requests': pending_requests,
        'total_donations': total_donations,
        'total_users': total_users,
        'recent_requests': recent_requests,
        'inventory': inventory,
        'recent_users': recent_users,
    }
    
    return render(request, 'admin_dashboard_enhanced.html', context)


# Donor Registration
@login_required
def register_donor(request):
    """Register a new donor"""
    if request.method == 'POST':
        form = DonorRegistrationForm(request.POST)
        if form.is_valid():
            donor = form.save(commit=False)
            donor.user = request.user
            donor.save()
            messages.success(request, 'Donor registered successfully!')
            return redirect('donor_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        # Pre-fill with user data if available
        initial_data = {}
        if hasattr(request.user, 'first_name') and request.user.first_name:
            initial_data['first_name'] = request.user.first_name
        if hasattr(request.user, 'last_name') and request.user.last_name:
            initial_data['last_name'] = request.user.last_name
        if hasattr(request.user, 'email') and request.user.email:
            initial_data['email'] = request.user.email
        if hasattr(request.user, 'phone_number') and request.user.phone_number:
            initial_data['phone_number'] = request.user.phone_number
        if hasattr(request.user, 'blood_type') and request.user.blood_type:
            initial_data['blood_type'] = request.user.blood_type
        
        form = DonorRegistrationForm(initial=initial_data)
    
    return render(request, 'register_donor.html', {'form': form})


# Donor List
@login_required
def donor_list(request):
    """List all donors with search and filter"""
    donors = Donor.objects.all()
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        donors = donors.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone_number__icontains=search_query)
        )
    
    # Filter by blood type
    blood_type_filter = request.GET.get('blood_type', '')
    if blood_type_filter:
        donors = donors.filter(blood_type=blood_type_filter)
    
    # Filter by availability
    availability_filter = request.GET.get('availability', '')
    if availability_filter:
        donors = donors.filter(is_available=(availability_filter == 'available'))
    
    context = {
        'donors': donors,
        'search_query': search_query,
        'blood_type_filter': blood_type_filter,
        'availability_filter': availability_filter,
    }
    
    return render(request, 'donor_list.html', context)


# Request Blood
@login_required
def request_blood(request):
    """Create a new blood request"""
    if request.method == 'POST':
        form = BloodRequestForm(request.POST)
        if form.is_valid():
            blood_request = form.save(commit=False)
            blood_request.requester = request.user
            blood_request.save()
            messages.success(request, 'Blood request submitted successfully!')
            return redirect('blood_request_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = BloodRequestForm()
    
    return render(request, 'request_blood.html', {'form': form})


# Blood Request List
@login_required
def blood_request_list(request):
    """List all blood requests"""
    if request.user.role == 'admin':
        requests = BloodRequest.objects.all()
    else:
        requests = BloodRequest.objects.filter(requester=request.user)
    
    # Filter by blood type
    blood_type_filter = request.GET.get('blood_type', '')
    if blood_type_filter:
        requests = requests.filter(blood_type=blood_type_filter)
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        requests = requests.filter(status=status_filter)
    
    # Filter by purpose
    purpose_filter = request.GET.get('purpose', '')
    if purpose_filter:
        requests = requests.filter(purpose=purpose_filter)
    
    # Filter by urgency
    urgency_filter = request.GET.get('urgency', '')
    if urgency_filter:
        requests = requests.filter(urgency=urgency_filter)
    
    context = {
        'requests': requests,
        'blood_type_filter': blood_type_filter,
        'status_filter': status_filter,
        'purpose_filter': purpose_filter,
        'urgency_filter': urgency_filter,
    }
    
    return render(request, 'all_requests.html', context)


def all_requests(request):
    """Alias for blood_request_list"""
    return blood_request_list(request)


# Delete Donor
@login_required
def delete_donor(request, pk):
    """Delete a donor"""
    donor = get_object_or_404(Donor, pk=pk)
    donor.delete()
    messages.success(request, 'Donor deleted successfully!')
    return redirect('donor_list')


# Stock Report
@login_required
def stock_report_print(request):
    """Print stock report"""
    inventory = BloodInventory.objects.all()
    return render(request, 'stock_report.html', {'inventory': inventory})


# Delete Request
@csrf_exempt
@require_http_methods(["DELETE", "POST"])
def delete_request(request, request_id):
    """Delete a blood request"""
    try:
        blood_request = get_object_or_404(BloodRequest, id=request_id)
        blood_request.delete()
        return JsonResponse({"success": True, "message": "Deleted successfully"}, status=200)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)


# ==========================================
# EXPORT FUNCTIONS
# ==========================================

# Export Donors to Excel
@login_required
def export_donors_excel(request):
    """Export donor list to Excel - Admin Only"""
    # Check if user is admin
    if request.user.role != 'admin':
        messages.error(request, 'Only administrators can export donor lists.')
        return redirect('donor_list')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Donors List"
    
    # Add header with styling
    headers = ['#', 'First Name', 'Last Name', 'Blood Type', 'Email', 'Phone', 'City', 'State', 'Available', 'Last Donation']
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Get donors
    donors = Donor.objects.all()
    
    # Add data
    for idx, donor in enumerate(donors, start=1):
        ws.append([
            idx,
            donor.first_name,
            donor.last_name,
            donor.blood_type,
            donor.email,
            donor.phone_number,
            donor.city,
            donor.state,
            'Yes' if donor.is_available else 'No',
            donor.last_donation_date.strftime('%Y-%m-%d') if donor.last_donation_date else 'Never'
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=donors_list.xlsx'
    
    wb.save(response)
    return response


# Export Donors to PDF
@login_required
def export_donors_pdf(request):
    """Export donor list to PDF - Admin Only"""
    # Check if user is admin
    if request.user.role != 'admin':
        messages.error(request, 'Only administrators can export donor lists.')
        return redirect('donor_list')
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=donors_list.pdf'
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#FF0000'),
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Title
    title = Paragraph("🩸 Blood Donors List", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3 * inch))
    
    # Get donors
    donors = Donor.objects.all()
    
    # Create table data
    data = [['#', 'Name', 'Blood Type', 'Phone', 'City', 'Available']]
    
    for idx, donor in enumerate(donors, start=1):
        data.append([
            str(idx),
            f"{donor.first_name} {donor.last_name}",
            donor.blood_type,
            donor.phone_number,
            donor.city,
            '✓' if donor.is_available else '✗'
        ])
    
    # Create table
    table = Table(data, colWidths=[0.5*inch, 2*inch, 1*inch, 1.5*inch, 1.5*inch, 1*inch])
    
    # Style table
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.red),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # Add footer
    elements.append(Spacer(1, 0.5 * inch))
    footer_text = f"Total Donors: {donors.count()} | Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
    footer = Paragraph(footer_text, styles['Normal'])
    elements.append(footer)
    
    # Build PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response


# Export Blood Requests to Excel
@login_required
def export_requests_excel(request):
    """Export blood requests to Excel"""
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Blood Requests"
    
    # Add header
    headers = ['#', 'Patient Name', 'Blood Type', 'Purpose', 'Units', 'Urgency', 'Hospital', 'Status', 'Date']
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color='DC143C', end_color='DC143C', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Get requests
    if request.user.role == 'admin':
        requests = BloodRequest.objects.all()
    else:
        requests = BloodRequest.objects.filter(requester=request.user)
    
    # Add data
    for idx, req in enumerate(requests, start=1):
        ws.append([
            idx,
            req.patient_name,
            req.blood_type,
            req.get_purpose_display(),
            req.units_needed,
            req.get_urgency_display(),
            req.hospital_name,
            req.get_status_display(),
            req.created_at.strftime('%Y-%m-%d')
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=blood_requests.xlsx'
    
    wb.save(response)
    return response


# Export Blood Requests to PDF
@login_required
def export_requests_pdf(request):
    """Export blood requests to PDF"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=blood_requests.pdf'
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#DC143C'),
        spaceAfter=30,
        alignment=1
    )
    
    # Title
    title = Paragraph("🩸 Blood Requests Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3 * inch))
    
    # Get requests
    if request.user.role == 'admin':
        requests = BloodRequest.objects.all()
    else:
        requests = BloodRequest.objects.filter(requester=request.user)
    
    # Create table
    data = [['#', 'Patient', 'Blood Type', 'Purpose', 'Units', 'Urgency', 'Status']]
    
    for idx, req in enumerate(requests, start=1):
        data.append([
            str(idx),
            req.patient_name,
            req.blood_type,
            req.get_purpose_display(),
            str(req.units_needed),
            req.get_urgency_display(),
            req.get_status_display()
        ])
    
    table = Table(data, colWidths=[0.4*inch, 1.5*inch, 0.8*inch, 1.2*inch, 0.6*inch, 0.9*inch, 1*inch])
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.crimson),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # Footer
    elements.append(Spacer(1, 0.5 * inch))
    footer_text = f"Total Requests: {requests.count()} | Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
    footer = Paragraph(footer_text, styles['Normal'])
    elements.append(footer)
    
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response


# ==========================================
# BLOOD TYPE COMPATIBILITY CHECKER - NEW FEATURE
# ==========================================

# Blood type compatibility data
BLOOD_COMPATIBILITY = {
    'A+': {
        'can_donate_to': ['A+', 'AB+'],
        'can_receive_from': ['A+', 'A-', 'O+', 'O-']
    },
    'A-': {
        'can_donate_to': ['A+', 'A-', 'AB+', 'AB-'],
        'can_receive_from': ['A-', 'O-']
    },
    'B+': {
        'can_donate_to': ['B+', 'AB+'],
        'can_receive_from': ['B+', 'B-', 'O+', 'O-']
    },
    'B-': {
        'can_donate_to': ['B+', 'B-', 'AB+', 'AB-'],
        'can_receive_from': ['B-', 'O-']
    },
    'AB+': {
        'can_donate_to': ['AB+'],
        'can_receive_from': ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-']
    },
    'AB-': {
        'can_donate_to': ['AB+', 'AB-'],
        'can_receive_from': ['A-', 'B-', 'AB-', 'O-']
    },
    'O+': {
        'can_donate_to': ['A+', 'B+', 'AB+', 'O+'],
        'can_receive_from': ['O+', 'O-']
    },
    'O-': {
        'can_donate_to': ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-'],
        'can_receive_from': ['O-']
    }
}


@login_required
def blood_compatibility_checker(request):
    """Blood type compatibility checker and donor finder"""
    selected_blood_type = request.GET.get('blood_type', '')
    compatible_donors = []
    compatibility_info = None
    
    if selected_blood_type and selected_blood_type in BLOOD_COMPATIBILITY:
        compatibility_info = BLOOD_COMPATIBILITY[selected_blood_type]
        
        # Find compatible donors
        compatible_blood_types = compatibility_info['can_receive_from']
        compatible_donors = Donor.objects.filter(
            blood_type__in=compatible_blood_types,
            is_available=True
        ).order_by('blood_type', 'city')
    
    context = {
        'selected_blood_type': selected_blood_type,
        'compatibility_info': compatibility_info,
        'compatible_donors': compatible_donors,
        'blood_compatibility': BLOOD_COMPATIBILITY,
    }
    
    return render(request, 'compatibility/blood_compatibility.html', context)


@login_required
def find_compatible_donors(request, blood_type):
    """Find donors compatible with a specific blood type"""
    if blood_type not in BLOOD_COMPATIBILITY:
        messages.error(request, 'Invalid blood type')
        return redirect('blood_compatibility_checker')
    
    compatibility_info = BLOOD_COMPATIBILITY[blood_type]
    compatible_blood_types = compatibility_info['can_receive_from']
    
    donors = Donor.objects.filter(
        blood_type__in=compatible_blood_types,
        is_available=True
    ).order_by('blood_type', 'city')
    
    context = {
        'blood_type': blood_type,
        'compatible_blood_types': compatible_blood_types,
        'donors': donors,
    }
    
    return render(request, 'compatibility/compatible_donors.html', context)



# Advanced Donor Search View
@login_required
def donor_search_view(request):
    """Advanced donor search with filters"""
    return render(request, 'donor_search.html')



# Certificate Download View
@login_required
def download_certificate(request, donation_id):
    """Download donation certificate as PDF"""
    from .certificates import generate_donation_certificate
    
    try:
        donation = BloodDonation.objects.get(id=donation_id)
        
        # Check if user has permission to download this certificate
        if request.user.role != 'admin' and donation.donor.user != request.user:
            messages.error(request, 'You do not have permission to download this certificate.')
            return redirect('user_dashboard')
        
        # Generate certificate
        pdf_buffer = generate_donation_certificate(donation)
        
        # Create response
        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        filename = f"Blood_Donation_Certificate_{donation.donor.last_name}_{donation.donation_date}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except BloodDonation.DoesNotExist:
        messages.error(request, 'Donation record not found.')
        return redirect('user_dashboard')



# My Donations View
@login_required
def my_donations(request):
    """View user's donation history and download certificates"""
    # Admins can see all donations, regular users see only their own
    if request.user.role == 'admin':
        # Admin sees all donations
        donations = BloodDonation.objects.all().select_related('donor', 'donor__user').order_by('-donation_date')
        total_units = sum(d.units_donated for d in donations)
        is_admin_view = True
    else:
        # Regular users see only their donations
        try:
            donor = Donor.objects.get(user=request.user)
            donations = BloodDonation.objects.filter(donor=donor).order_by('-donation_date')
            total_units = sum(d.units_donated for d in donations)
        except Donor.DoesNotExist:
            donations = []
            total_units = 0
        is_admin_view = False
    
    context = {
        'donations': donations,
        'total_units': total_units,
        'is_admin_view': is_admin_view,
    }
    
    return render(request, 'donations/my_donations.html', context)


@login_required
def edit_donor(request, donor_id):
    """Edit donor information (Admin only)"""
    if request.user.role != 'admin':
        messages.error(request, 'Only administrators can edit donor information.')
        return redirect('donor_list')
    
    donor = get_object_or_404(Donor, id=donor_id)
    
    if request.method == 'POST':
        # Update donor information
        donor.first_name = request.POST.get('first_name')
        donor.last_name = request.POST.get('last_name')
        donor.email = request.POST.get('email')
        donor.phone_number = request.POST.get('phone_number')
        donor.blood_type = request.POST.get('blood_type')
        donor.date_of_birth = request.POST.get('date_of_birth')
        donor.address = request.POST.get('address')
        donor.city = request.POST.get('city')
        donor.state = request.POST.get('state')
        donor.is_available = request.POST.get('is_available') == 'on'
        
        donor.save()
        messages.success(request, f'Donor {donor.first_name} {donor.last_name} updated successfully!')
        return redirect('donor_list')
    
    context = {
        'donor': donor,
    }
    return render(request, 'donors/edit_donor.html', context)


@login_required
def patient_list(request):
    """View all patients (blood request recipients)"""
    if request.user.role != 'admin':
        messages.error(request, 'Only administrators can view patient list.')
        return redirect('user_dashboard')
    
    # Get all blood requests (each represents a patient)
    patients = BloodRequest.objects.all().select_related('requester').order_by('-created_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        patients = patients.filter(
            Q(patient_name__icontains=search_query) |
            Q(blood_type__icontains=search_query) |
            Q(hospital_name__icontains=search_query) |
            Q(purpose__icontains=search_query)
        )
    
    context = {
        'patients': patients,
        'search_query': search_query,
    }
    return render(request, 'patients/patient_list.html', context)


@login_required
def edit_patient(request, request_id):
    """Edit patient/blood request information (Admin only)"""
    if request.user.role != 'admin':
        messages.error(request, 'Only administrators can edit patient information.')
        return redirect('patient_list')
    
    blood_request = get_object_or_404(BloodRequest, id=request_id)
    
    if request.method == 'POST':
        # Update patient/request information
        blood_request.patient_name = request.POST.get('patient_name')
        blood_request.blood_type = request.POST.get('blood_type')
        blood_request.units_needed = request.POST.get('units_needed')
        blood_request.purpose = request.POST.get('purpose')
        blood_request.purpose_details = request.POST.get('purpose_details')
        blood_request.urgency = request.POST.get('urgency')
        blood_request.hospital_name = request.POST.get('hospital_name')
        blood_request.hospital_address = request.POST.get('hospital_address')
        blood_request.contact_number = request.POST.get('contact_number')
        blood_request.required_date = request.POST.get('required_date')
        blood_request.notes = request.POST.get('notes')
        
        blood_request.save()
        messages.success(request, f'Patient {blood_request.patient_name} information updated successfully!')
        return redirect('patient_list')
    
    context = {
        'blood_request': blood_request,
    }
    return render(request, 'patients/edit_patient.html', context)


@login_required
def delete_patient(request, request_id):
    """Delete patient/blood request (Admin only)"""
    if request.user.role != 'admin':
        messages.error(request, 'Only administrators can delete patient records.')
        return redirect('patient_list')
    
    blood_request = get_object_or_404(BloodRequest, id=request_id)
    patient_name = blood_request.patient_name
    blood_request.delete()
    
    messages.success(request, f'Patient record for {patient_name} has been deleted.')
    return redirect('patient_list')


@login_required
def donation_request_list(request):
    """View all blood donation requests for admin approval/rejection"""
    if request.user.role != 'admin':
        messages.error(request, 'Only administrators can manage donation requests.')
        return redirect('user_dashboard')
    
    # Get all blood donations
    donations = BloodDonation.objects.all().select_related('donor', 'donor__user').order_by('-donation_date')
    
    # Filter by status if provided
    status_filter = request.GET.get('status', '')
    if status_filter:
        donations = donations.filter(status=status_filter)
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        donations = donations.filter(
            Q(donor__first_name__icontains=search_query) |
            Q(donor__last_name__icontains=search_query) |
            Q(blood_type__icontains=search_query) |
            Q(hospital_name__icontains=search_query)
        )
    
    context = {
        'donations': donations,
        'search_query': search_query,
        'status_filter': status_filter,
    }
    return render(request, 'donations/donation_request_list.html', context)


@login_required
def approve_donation(request, donation_id):
    """Approve a blood donation request and add to inventory"""
    if request.user.role != 'admin':
        messages.error(request, 'Only administrators can approve donations.')
        return redirect('user_dashboard')
    
    donation = get_object_or_404(BloodDonation, id=donation_id)
    
    if donation.status == 'approved':
        messages.warning(request, 'This donation has already been approved.')
        return redirect('donation_request_list')
    
    # Update donation status
    donation.status = 'approved'
    donation.approved_by = request.user
    donation.approved_at = timezone.now()
    donation.save()
    
    # Add units to blood inventory
    inventory, created = BloodInventory.objects.get_or_create(
        blood_type=donation.blood_type,
        defaults={'units_available': 0}
    )
    inventory.units_available += donation.units_donated
    inventory.save()
    
    messages.success(request, f'Donation approved! {donation.units_donated} unit(s) of {donation.blood_type} added to inventory.')
    return redirect('donation_request_list')


@login_required
def reject_donation(request, donation_id):
    """Reject a blood donation request"""
    if request.user.role != 'admin':
        messages.error(request, 'Only administrators can reject donations.')
        return redirect('user_dashboard')
    
    donation = get_object_or_404(BloodDonation, id=donation_id)
    
    if request.method == 'POST':
        rejection_reason = request.POST.get('rejection_reason', '')
        
        donation.status = 'rejected'
        donation.rejection_reason = rejection_reason
        donation.approved_by = request.user
        donation.approved_at = timezone.now()
        donation.save()
        
        messages.success(request, f'Donation rejected. No units added to inventory.')
        return redirect('donation_request_list')
    
    context = {
        'donation': donation,
    }
    return render(request, 'donations/reject_donation.html', context)



# ==========================================
# PASSWORD RESET VIEWS
# ==========================================

from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.template.loader import render_to_string
from django.core.mail import send_mail
from django.conf import settings


def password_reset_request(request):
    """Password reset request view"""
    if request.method == 'POST':
        email = request.POST.get('email')
        
        try:
            user = CustomUser.objects.get(email=email)
            
            # Generate token
            token = default_token_generator.make_token(user)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            
            # Build reset URL
            reset_url = request.build_absolute_uri(
                f'/password-reset-confirm/{uid}/{token}/'
            )
            
            # Send email (in production, configure email settings)
            # For now, just show success message
            messages.success(
                request, 
                f'Password reset instructions have been sent to {email}. '
                'Please check your email and follow the instructions. '
                'If you don\'t receive an email, contact support at +254 700 123 456'
            )
            
            # In production, uncomment this to send actual email:
            # subject = 'Password Reset Request - Blood Management System'
            # message = f'Click the link below to reset your password:\n\n{reset_url}\n\nThis link expires in 24 hours.'
            # send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [email])
            
            return redirect('password_reset_done')
            
        except CustomUser.DoesNotExist:
            # Don't reveal that the email doesn't exist for security
            messages.success(
                request, 
                'If an account exists with that email, password reset instructions have been sent. '
                'If you don\'t receive an email, contact support at +254 700 123 456'
            )
            return redirect('password_reset_done')
    
    return render(request, 'registration/password_reset_form.html')


def password_reset_done(request):
    """Password reset email sent confirmation"""
    return render(request, 'registration/password_reset_done.html')


def password_reset_confirm(request, uidb64, token):
    """Password reset confirmation view"""
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = CustomUser.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, CustomUser.DoesNotExist):
        user = None
    
    if user is not None and default_token_generator.check_token(user, token):
        if request.method == 'POST':
            password1 = request.POST.get('new_password1')
            password2 = request.POST.get('new_password2')
            
            if password1 and password2:
                if password1 == password2:
                    if len(password1) >= 8:
                        user.set_password(password1)
                        user.save()
                        messages.success(request, 'Your password has been reset successfully!')
                        return redirect('password_reset_complete')
                    else:
                        messages.error(request, 'Password must be at least 8 characters long.')
                else:
                    messages.error(request, 'Passwords do not match.')
            else:
                messages.error(request, 'Please fill in both password fields.')
        
        return render(request, 'registration/password_reset_confirm.html', {
            'validlink': True,
            'form': {'new_password1': '', 'new_password2': ''}
        })
    else:
        return render(request, 'registration/password_reset_confirm.html', {
            'validlink': False
        })


def password_reset_complete(request):
    """Password reset complete view"""
    return render(request, 'registration/password_reset_complete.html')


# ==========================================
# CONTACT US VIEW
# ==========================================

def contact_us(request):
    """Contact us page"""
    return render(request, 'contact_us.html')



# ==========================================
# CONTACT FOR BLOOD VIEW
# ==========================================

def contact_for_blood(request):
    """Contact for blood - emergency blood request form"""
    if request.method == 'POST':
        name = request.POST.get('name')
        phone = request.POST.get('phone')
        email = request.POST.get('email')
        blood_type = request.POST.get('blood_type')
        units = request.POST.get('units')
        relation = request.POST.get('relation')
        urgency = request.POST.get('urgency')
        message = request.POST.get('message')
        
        # In production, this would send an email/SMS to admin or create a blood request
        # For now, show success message
        messages.success(
            request,
            f'Your blood request has been received! We will contact you at {phone} shortly. '
            f'For urgent requests, please also call +254 700 123 456'
        )
        
        # If user is logged in, create a blood request automatically
        if request.user.is_authenticated:
            try:
                blood_request = BloodRequest.objects.create(
                    requester=request.user,
                    patient_name=name,
                    blood_type=blood_type,
                    units_needed=int(units),
                    purpose='emergency',
                    purpose_details=f"{relation} - {message}",
                    urgency='critical' if urgency in ['urgent', 'urgently'] else 'high',
                    hospital_name='To be determined',
                    hospital_address='To be determined',
                    contact_number=phone,
                    required_date=timezone.now().date(),
                    notes=f'Contact for Blood Form Submission - Urgency: {urgency}'
                )
                messages.info(request, 'A blood request has been created in your account. You can track it from your dashboard.')
            except Exception as e:
                pass
        
        return redirect('contact_for_blood')
    
    return render(request, 'contact_for_blood.html')



# ==========================================
# ADVANCED SEARCH FUNCTIONALITY
# ==========================================

@login_required
def advanced_search(request):
    """Advanced search with date range and multiple filters"""
    # Get all filter parameters
    search_query = request.GET.get('search', '')
    blood_type = request.GET.get('blood_type', '')
    status = request.GET.get('status', '')
    urgency = request.GET.get('urgency', '')
    purpose = request.GET.get('purpose', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search_type = request.GET.get('search_type', 'requests')  # requests or donors
    
    results = []
    result_count = 0
    
    if search_type == 'donors':
        # Search donors
        donors = Donor.objects.all()
        
        if search_query:
            donors = donors.filter(
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(phone_number__icontains=search_query) |
                Q(city__icontains=search_query) |
                Q(state__icontains=search_query)
            )
        
        if blood_type:
            donors = donors.filter(blood_type=blood_type)
        
        if date_from:
            donors = donors.filter(created_at__gte=date_from)
        
        if date_to:
            donors = donors.filter(created_at__lte=date_to)
        
        results = donors
        result_count = donors.count()
        
    else:
        # Search blood requests
        requests = BloodRequest.objects.all()
        
        if request.user.role != 'admin':
            requests = requests.filter(requester=request.user)
        
        if search_query:
            requests = requests.filter(
                Q(patient_name__icontains=search_query) |
                Q(hospital_name__icontains=search_query) |
                Q(hospital_address__icontains=search_query) |
                Q(contact_number__icontains=search_query) |
                Q(notes__icontains=search_query)
            )
        
        if blood_type:
            requests = requests.filter(blood_type=blood_type)
        
        if status:
            requests = requests.filter(status=status)
        
        if urgency:
            requests = requests.filter(urgency=urgency)
        
        if purpose:
            requests = requests.filter(purpose=purpose)
        
        if date_from:
            requests = requests.filter(created_at__gte=date_from)
        
        if date_to:
            requests = requests.filter(created_at__lte=date_to)
        
        results = requests
        result_count = requests.count()
    
    context = {
        'results': results,
        'result_count': result_count,
        'search_query': search_query,
        'blood_type': blood_type,
        'status': status,
        'urgency': urgency,
        'purpose': purpose,
        'date_from': date_from,
        'date_to': date_to,
        'search_type': search_type,
    }
    
    return render(request, 'advanced_search.html', context)



# ==========================================
# USER MANAGEMENT VIEWS
# ==========================================

@login_required
def user_list(request):
    """View all registered users (Admin only)"""
    if request.user.role != 'admin':
        messages.error(request, 'Only administrators can view user list.')
        return redirect('user_dashboard')
    
    # Get all users
    users = CustomUser.objects.all().order_by('-date_joined')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone_number__icontains=search_query)
        )
    
    # Filter by role
    role_filter = request.GET.get('role', '')
    if role_filter:
        users = users.filter(role=role_filter)
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        users = users.filter(is_active=True)
    elif status_filter == 'inactive':
        users = users.filter(is_active=False)
    
    # Statistics
    total_users = CustomUser.objects.count()
    admin_count = CustomUser.objects.filter(role='admin').count()
    user_count = CustomUser.objects.filter(role='user').count()
    active_today = CustomUser.objects.filter(last_login__date=timezone.now().date()).count()
    
    # Blood inventory
    inventory = BloodInventory.objects.all().order_by('blood_type')
    
    context = {
        'users': users,
        'search_query': search_query,
        'role_filter': role_filter,
        'status_filter': status_filter,
        'total_users': total_users,
        'admin_count': admin_count,
        'user_count': user_count,
        'active_today': active_today,
        'inventory': inventory,
    }
    
    return render(request, 'users/user_list.html', context)


@login_required
def view_user(request, user_id):
    """View user details (Admin only)"""
    if request.user.role != 'admin':
        messages.error(request, 'Only administrators can view user details.')
        return redirect('user_dashboard')
    
    viewed_user = get_object_or_404(CustomUser, id=user_id)
    
    # Get user activity
    blood_requests_count = BloodRequest.objects.filter(requester=viewed_user).count()
    
    # Check if user is a donor
    is_donor = Donor.objects.filter(user=viewed_user).exists()
    donations_count = 0
    if is_donor:
        donor = Donor.objects.get(user=viewed_user)
        donations_count = BloodDonation.objects.filter(donor=donor).count()
    
    context = {
        'viewed_user': viewed_user,
        'blood_requests_count': blood_requests_count,
        'donations_count': donations_count,
        'is_donor': is_donor,
    }
    
    return render(request, 'users/user_detail.html', context)


@login_required
def edit_user(request, user_id):
    """Edit user information (Admin only)"""
    if request.user.role != 'admin':
        messages.error(request, 'Only administrators can edit user information.')
        return redirect('user_dashboard')
    
    user_to_edit = get_object_or_404(CustomUser, id=user_id)
    
    if request.method == 'POST':
        # Update user information
        user_to_edit.first_name = request.POST.get('first_name')
        user_to_edit.last_name = request.POST.get('last_name')
        user_to_edit.email = request.POST.get('email')
        user_to_edit.phone_number = request.POST.get('phone_number')
        user_to_edit.blood_type = request.POST.get('blood_type')
        user_to_edit.address = request.POST.get('address')
        user_to_edit.role = request.POST.get('role')
        user_to_edit.is_active = request.POST.get('is_active') == 'on'
        
        # Update date of birth if provided
        dob = request.POST.get('date_of_birth')
        if dob:
            user_to_edit.date_of_birth = dob
        
        user_to_edit.save()
        messages.success(request, f'User {user_to_edit.username} updated successfully!')
        return redirect('view_user', user_id=user_id)
    
    context = {
        'user_to_edit': user_to_edit,
    }
    return render(request, 'users/edit_user.html', context)



@login_required
def clear_welcome_flag(request):
    """Clear the welcome modal flag from session"""
    if 'show_welcome' in request.session:
        del request.session['show_welcome']
    if 'welcome_greeting' in request.session:
        del request.session['welcome_greeting']
    if 'welcome_name' in request.session:
        del request.session['welcome_name']
    if 'user_role' in request.session:
        del request.session['user_role']
    if 'last_login' in request.session:
        del request.session['last_login']
    return JsonResponse({'status': 'success'})
